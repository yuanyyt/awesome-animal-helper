"""Read animal names from the source workbook."""

from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook


def read_animals(workbook_path: Path) -> list[str]:
    """Read unique animal names from column B, preserving first-seen order."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        animals: list[str] = []
        seen: set[str] = set()
        for (value,) in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            for animal in _split_animals(value):
                if animal not in seen:
                    seen.add(animal)
                    animals.append(animal)
        return animals
    finally:
        workbook.close()


def _split_animals(value: object) -> Iterable[str]:
    if value is None:
        return
    for animal in str(value).split():
        normalized = animal.strip()
        if normalized:
            yield normalized
