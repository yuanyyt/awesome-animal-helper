"""Load and query the local animal CSV and venue workbook."""

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from src.backend.domain.models import AnimalDetail, AnimalListResponse, SiteSummary

DATA_DIR = Path(__file__).resolve().parents[2] / "data"
ANIMALS_PATH = DATA_DIR / "animals.csv"
SITES_PATH = DATA_DIR / "animal_sites.xlsx"

_REQUIRED_COLUMNS = {
    "动物",
    "学名",
    "分类",
    "栖息地",
    "分布",
    "食性",
    "行为",
    "繁殖",
    "保护状态",
    "趣味事实",
    "来源URL",
    "语言",
    "状态",
}


class AnimalRepository:
    """Immutable in-memory view of the two source files."""

    def __init__(
        self,
        animals_path: Path = ANIMALS_PATH,
        sites_path: Path = SITES_PATH,
    ) -> None:
        site_order, animal_sites = _read_site_assignments(sites_path)
        self._animals = _read_animals(animals_path, animal_sites)
        self._site_order = site_order

    def query(
        self,
        *,
        q: str | None = None,
        site: str | None = None,
        name: str | None = None,
    ) -> AnimalListResponse:
        """Filter animals while preserving the source CSV order."""

        query_text = _clean_filter(q).casefold()
        site_name = _clean_filter(site)
        animal_name = _clean_filter(name)

        items = [
            animal
            for animal in self._animals
            if (not animal_name or animal.name == animal_name)
            and (not site_name or site_name in animal.sites)
            and (not query_text or _matches_query(animal, query_text))
        ]
        return AnimalListResponse(
            items=items,
            sites=self._site_summaries(),
            total=len(self._animals),
            filtered_count=len(items),
        )

    def site_summaries(self) -> list[SiteSummary]:
        """Return venue metadata in workbook order."""

        return self._site_summaries()

    def _site_summaries(self) -> list[SiteSummary]:
        counts = Counter(site for animal in self._animals for site in animal.sites)
        return [
            SiteSummary(name=site, animal_count=counts[site])
            for site in self._site_order
        ]


def _read_site_assignments(path: Path) -> tuple[list[str], dict[str, list[str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        site_order: list[str] = []
        animal_sites: dict[str, list[str]] = defaultdict(list)
        for site_value, animals_value in sheet.iter_rows(min_row=2, values_only=True):
            site = str(site_value or "").strip()
            if not site:
                continue
            site_order.append(site)
            for animal in str(animals_value or "").split():
                if site not in animal_sites[animal]:
                    animal_sites[animal].append(site)
        return site_order, dict(animal_sites)
    finally:
        workbook.close()


def _read_animals(
    path: Path,
    animal_sites: dict[str, list[str]],
) -> list[AnimalDetail]:
    with path.open(encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        columns = set(reader.fieldnames or ())
        missing = sorted(_REQUIRED_COLUMNS - columns)
        if missing:
            raise ValueError(f"animals.csv 缺少字段：{'、'.join(missing)}")

        animals: list[AnimalDetail] = []
        seen_names: set[str] = set()
        for row in reader:
            name = row["动物"].strip()
            if not name or name in seen_names:
                continue
            seen_names.add(name)
            animals.append(
                AnimalDetail(
                    name=name,
                    scientific_name=_optional(row["学名"]),
                    taxonomy=_optional(row["分类"]),
                    habitat=_optional(row["栖息地"]),
                    distribution=_optional(row["分布"]),
                    diet=_optional(row["食性"]),
                    behavior=_optional(row["行为"]),
                    reproduction=_optional(row["繁殖"]),
                    conservation_status=_optional(row["保护状态"]),
                    fun_facts=_split_facts(row["趣味事实"]),
                    source_url=_optional(row["来源URL"]),
                    language=_optional(row["语言"]),
                    data_status=row["状态"].strip() or "partial",
                    sites=animal_sites.get(name, []),
                )
            )
        return animals


def _matches_query(animal: AnimalDetail, query: str) -> bool:
    searchable = " ".join(
        value for value in (animal.name, animal.scientific_name, animal.taxonomy) if value
    )
    return query in searchable.casefold()


def _clean_filter(value: str | None) -> str:
    return (value or "").strip()


def _optional(value: str | None) -> str | None:
    cleaned = (value or "").strip()
    return cleaned or None


def _split_facts(value: str | None) -> list[str]:
    text = (value or "").strip()
    if not text:
        return []
    return [fact.strip() for fact in re.split(r"\s*；\s*", text) if fact.strip()]
